#!/usr/bin/env python3
"""Burndown report generator for P0/P1/P2 vulnerability data."""

import os
import sys
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURATION — update file paths and labels before running
# ==============================================================================

# Input CSV files — today's vulnerability snapshot
TODAY_P0_FILE = "data/today_p0.csv"
TODAY_P1_FILE = "data/today_p1.csv"
TODAY_P2_FILE = "data/today_p2.csv"

# Input CSV files — baseline snapshot (start-of-week or program baseline)
BASELINE_P0_FILE = "data/baseline_p0.csv"
BASELINE_P1_FILE = "data/baseline_p1.csv"
BASELINE_P2_FILE = "data/baseline_p2.csv"

# Output workbook path
OUTPUT_FILE = f"burndown_report_{datetime.now().strftime('%Y%m%d')}.xlsx"

# Normalized column names — must match the headers in the P0 CSV files exactly
INSTANCE_ID_COL         = "Instance ID"
DEPLOYMENT_LOCATION_COL = "Deployment Location"
MC2_COL                 = "MC2"

# P1 / P2 raw export field names → normalized column names used by P0
# P0 files are expected to already carry the normalized headers below.
COLUMN_MAPPING = {
    "saltminer.attributes.issue_instance_id":                    INSTANCE_ID_COL,
    "saltminer.asset.attributes.deployment_location":            DEPLOYMENT_LOCATION_COL,
    "saltminer.inventory_asset.attributes.appmap.cio":           "CIO",
    "vulnerability.severity":                                    "Severity",
    "vulnerability.name":                                        "Vulnerability Name",
    "saltminer.asset.version":                                   "Release Version",
    "saltminer.inventory_asset.attributes.appmap.apm_number":   "App ID",
    "saltminer.inventory_asset.attributes.appmap.application_name": "App Name",
    "saltminer.inventory_asset.attributes.appmap.mc2":           MC2_COL,
}

# Report header labels
MC2_HEADER_LABEL  = "MC-2"
GRAND_TOTAL_LABEL = "Grand Total"

# Date values computed once at startup
_today       = datetime.now()
_monday      = _today - timedelta(days=_today.weekday())  # most recent Monday
TODAY_LABEL  = _today.strftime("%m/%d/%Y")
MONDAY_LABEL = _monday.strftime("%m/%d/%Y")

# Column header text templates
STANDALONE_HEADER  = f"<<Start of\nthe Week>>\n{MONDAY_LABEL}"
LOCATION_HEADER_FMT = "Deployed at {loc} - {today}\n[Cumulative from Start of the week]"

# Excel color palette (6-digit RGB hex, no leading #)
COLORS = {
    "mc2_header":        "1F3864",  # dark navy — MC-2 col header + metric sub-headers
    "standalone_header": "9DC3E6",  # light periwinkle — start-of-week column header
    "location_header":   "E36C09",  # orange — per-location group header
    "mc2_data":          "F2F2F2",  # light gray — MC-2 label cells
    "alt_row":           "D9D9D9",  # alternating data row background
    "grand_total":       "BDD7EE",  # light blue — grand total row
    "white":             "FFFFFF",
    "black":             "000000",
}

# ==============================================================================
# END CONFIGURATION
# ==============================================================================

METRICS = ["Baseline", "Current", "Newly Added", "Closed", "Closure Rate"]

_THIN  = Side(style="thin")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Data loading ───────────────────────────────────────────────────────────────

def load_data(filepath):
    """Load a UTF-8 CSV file; every column is kept as text."""
    if not os.path.exists(filepath):
        sys.exit(f"ERROR: data file not found — {filepath}")
    return pd.read_csv(filepath, dtype=str, encoding="utf-8-sig").fillna("")


def normalize_columns(df):
    """Rename P1/P2 raw export headers to the normalized names used by P0."""
    return df.rename(columns=COLUMN_MAPPING)


# ── Burndown metric computation ────────────────────────────────────────────────

def _instance_ids(df, mc2=None, location=None):
    """Return the set of non-empty Instance IDs matching optional MC2 / location filters."""
    mask = pd.Series(True, index=df.index)
    if mc2 is not None:
        mask &= df[MC2_COL] == mc2
    if location is not None:
        mask &= df[DEPLOYMENT_LOCATION_COL] == location
    ids = df.loc[mask, INSTANCE_ID_COL]
    return set(ids[ids != ""])


def compute_metrics(today_df, baseline_df, mc2=None, location=None):
    """
    Compute the five burndown metrics for an optional MC2 / Deployment Location slice.

    Newly Added  = in today but not in baseline (matched by Instance ID)
    Closed       = in baseline but not in today  (matched by Instance ID)
    Closure Rate = Closed / Baseline count
    """
    baseline_ids = _instance_ids(baseline_df, mc2, location)
    today_ids    = _instance_ids(today_df,    mc2, location)
    baseline_n   = len(baseline_ids)
    closed       = len(baseline_ids - today_ids)
    return {
        "Baseline":     baseline_n,
        "Current":      len(today_ids),
        "Newly Added":  len(today_ids - baseline_ids),
        "Closed":       closed,
        "Closure Rate": closed / baseline_n if baseline_n else 0.0,
    }


# ── Report data structure ──────────────────────────────────────────────────────

def generate_report(today_df, baseline_df):
    """
    Build the burndown report DataFrame.

    Rows    — one per unique MC2 value found across both datasets, plus a Grand Total row.
    Columns — MC2 label, then for each Deployment Location:
                a standalone baseline key  ({loc}|standalone)
                and the five METRICS keys  ({loc}|<metric>).

    Returns (report_df, ordered list of deployment location strings).
    """
    def unique_sorted(col):
        return sorted(
            set(today_df[col].replace("", pd.NA).dropna()) |
            set(baseline_df[col].replace("", pd.NA).dropna())
        )

    mc2_values = unique_sorted(MC2_COL)
    locations  = unique_sorted(DEPLOYMENT_LOCATION_COL)

    def build_row(mc2_val, label):
        row = {MC2_COL: label}
        for loc in locations:
            m = compute_metrics(today_df, baseline_df, mc2=mc2_val, location=loc)
            row[f"{loc}|standalone"] = m["Baseline"]  # standalone col mirrors baseline
            for metric, value in m.items():
                row[f"{loc}|{metric}"] = value
        return row

    rows = [build_row(mc2, mc2) for mc2 in mc2_values]
    rows.append(build_row(None, GRAND_TOTAL_LABEL))  # mc2=None → no MC2 filter

    return pd.DataFrame(rows), locations


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size)


def _write(ws, row, col, value=None, fill_color=None, bold=False,
           font_color="000000", font_size=11, alignment=CENTER,
           border=BORDER, num_fmt=None):
    """Write value and formatting to a single cell."""
    c = ws.cell(row, col)
    if value is not None:
        c.value = value
    if fill_color:
        c.fill = _fill(fill_color)
    c.font      = _font(bold=bold, color=font_color, size=font_size)
    c.alignment = alignment
    c.border    = border
    if num_fmt:
        c.number_format = num_fmt
    return c


# ── Sheet writer ───────────────────────────────────────────────────────────────

def write_excel_sheet(ws, report_df, locations, label):
    """
    Write a fully formatted burndown sheet matching the standard report layout:

    Row 1  — workbook title banner
    Row 2  — MC-2 header (spans rows 2-3) | per-location: standalone header (spans 2-3)
               + location group header (row 2, orange)
    Row 3  — metric sub-headers under each location group (Baseline … Closure Rate)
    Row 4+ — data rows: one per MC2, Grand Total last
    """
    n_metrics    = len(METRICS)
    cols_per_loc = 1 + n_metrics          # 1 standalone + 5 metrics
    total_cols   = 1 + len(locations) * cols_per_loc

    # ── Row 1: title banner ───────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    _write(ws, 1, 1,
           value=f"{label} Vulnerability Burndown Report  ·  {_today.strftime('%B %d, %Y')}",
           fill_color=COLORS["mc2_header"], bold=True,
           font_color=COLORS["white"], font_size=14)
    ws.row_dimensions[1].height = 30

    # ── Row 2: MC-2 header (spans rows 2–3) ──────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    _write(ws, 2, 1, value=MC2_HEADER_LABEL,
           fill_color=COLORS["mc2_header"], bold=True,
           font_color=COLORS["white"])

    # ── Rows 2–3: location columns ────────────────────────────────────────────
    col = 2
    for loc in locations:
        # Standalone "Start of Week" column — spans both header rows, light blue
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        _write(ws, 2, col, value=STANDALONE_HEADER,
               fill_color=COLORS["standalone_header"], bold=True,
               font_color=COLORS["black"])
        col += 1

        # Location group header — row 2 only, orange, spans all metric sub-columns
        ws.merge_cells(start_row=2, start_column=col,
                       end_row=2, end_column=col + n_metrics - 1)
        _write(ws, 2, col,
               value=LOCATION_HEADER_FMT.format(loc=loc, today=TODAY_LABEL),
               fill_color=COLORS["location_header"], bold=True,
               font_color=COLORS["white"])

        # Metric sub-headers — row 3, dark navy
        for metric in METRICS:
            _write(ws, 3, col, value=metric,
                   fill_color=COLORS["mc2_header"], bold=True,
                   font_color=COLORS["white"])
            col += 1

    ws.row_dimensions[2].height = 44
    ws.row_dimensions[3].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, (_, row) in enumerate(report_df.iterrows()):
        r        = 4 + i
        is_total = row[MC2_COL] == GRAND_TOTAL_LABEL
        alt      = i % 2 == 1 and not is_total
        row_bg   = (COLORS["grand_total"] if is_total
                    else COLORS["alt_row"] if alt
                    else "FFFFFF")

        # MC-2 label cell
        _write(ws, r, 1, value=row[MC2_COL],
               fill_color=COLORS["grand_total"] if is_total else COLORS["mc2_data"],
               bold=is_total)

        col = 2
        for loc in locations:
            # Standalone baseline value (mirrors the Baseline metric)
            standalone = int(row.get(f"{loc}|standalone", 0) or 0)
            _write(ws, r, col, value=standalone, fill_color=row_bg, bold=is_total)
            col += 1

            # Five metric values
            for metric in METRICS:
                value = row.get(f"{loc}|{metric}", 0)
                if metric == "Closure Rate":
                    _write(ws, r, col, value=float(value),
                           fill_color=row_bg, bold=is_total, num_fmt="0.0%")
                else:
                    _write(ws, r, col, value=int(value) if value else 0,
                           fill_color=row_bg, bold=is_total)
                col += 1

        ws.row_dimensions[r].height = 16

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 28  # MC-2
    col = 2
    for _ in locations:
        ws.column_dimensions[get_column_letter(col)].width = 13   # standalone
        col += 1
        for metric in METRICS:
            ws.column_dimensions[get_column_letter(col)].width = (
                15 if metric in ("Newly Added", "Closure Rate") else 11
            )
            col += 1

    # Freeze header rows and MC-2 column so they stay visible while scrolling
    ws.freeze_panes = ws.cell(4, 2)


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    print("Loading data files...")
    today_p0    = load_data(TODAY_P0_FILE)
    today_p1    = load_data(TODAY_P1_FILE)
    today_p2    = load_data(TODAY_P2_FILE)
    baseline_p0 = load_data(BASELINE_P0_FILE)
    baseline_p1 = load_data(BASELINE_P1_FILE)
    baseline_p2 = load_data(BASELINE_P2_FILE)

    print("Normalizing P1 and P2 columns...")
    today_p1    = normalize_columns(today_p1)
    today_p2    = normalize_columns(today_p2)
    baseline_p1 = normalize_columns(baseline_p1)
    baseline_p2 = normalize_columns(baseline_p2)

    print("Building reports...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # discard the default blank sheet

    for label, today, baseline in [
        ("P0", today_p0, baseline_p0),
        ("P1", today_p1, baseline_p1),
        ("P2", today_p2, baseline_p2),
    ]:
        report_df, locations = generate_report(today, baseline)
        ws = wb.create_sheet(title=label)
        write_excel_sheet(ws, report_df, locations, label)
        mc2_count = len(report_df) - 1  # exclude Grand Total row
        print(f"  {label}: {mc2_count} MC-2 rows across {len(locations)} location(s)")

    wb.save(OUTPUT_FILE)
    print(f"\nReport saved → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
